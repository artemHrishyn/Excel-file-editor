import openpyxl

# читаем excel-файл
wb1 = openpyxl.load_workbook('PnK 2019.xlsx')
wb2 = openpyxl.load_workbook('PnK 2020.xlsx')

# печатаем список листов
sheets1 = wb1.sheetnames
sheets2 = wb2.sheetnames

for sheet1 in sheets1:
    print(sheet1)
for sheet2 in sheets2:
    print(sheet2)

# получаем активный лист
sheet1 = wb1.active

# печатаем значение ячейки A1
print(sheet1['A1'].value)
# печатаем значение ячейки B1
print(sheet1['B1'].value)

# получаем активный лист
sheet2 = wb2.active

# печатаем значение ячейки A1
print(sheet2['A1'].value)
# печатаем значение ячейки B1
print(sheet2['B1'].value)

rows = sheet1.max_row
cols = sheet1.max_column

def exel_2020(self, znach2):
    for cell2 in sheet2['B']:
        znach2 = cell2.value
    return znach2

for cell1 in sheet1['B']:
    znach1 = cell1.value
